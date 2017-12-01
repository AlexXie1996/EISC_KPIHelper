import os
import re
import random
from openpyxl.reader.excel import load_workbook
from borden import *

###################################################################
# :辅助函数                                                       #
#                                                                 #
# 函数                                                            #
#    - name_to_path_by_rstr 通过提供的字符串找对应表的路径        #
#    - name_to_path 通过提供的名字找对应表的路径                  #
#    - check_score_int 检查一个整数数字，可限定范围               #
#    - check_score_float 检查一个浮点数数字，可限定范围           #
#    - shuffle 打乱(评论的)顺序                                 　#
#    - find_dep_by_name 通过提供的名字找其对应的部门              #
###################################################################
def name_to_path_by_rstr(rstr, path, text):
	"""
	:指定任意一个字串，正则匹配返回含有这个字串的表（含检查）
	params：
		- rstr 要搜索的字符串，匹配项
		- path 要搜索的当前路径
		- text 断言的提示字段
	return:
		- cur_path[0] 匹配的第一个表名
	"""
	cur_path = [i for i in os.listdir(path) if re.search(rstr, i) is not None]
	assert len(cur_path) > 0, "在路径: {0} 中{1}".format(path, text)

	return cur_path[0]

def name_to_path(name, path):
	"""
	:指定一个名字，正则匹配返回含有这个名字的表（含检查）
	params：
		- name 要搜索的名字，匹配项
		- path 要搜索的当前路径
	return:
		- cur_path[0] 匹配的第一个表名	
	"""
	cur_path = [i for i in os.listdir(path) if re.search(name, i) is not None]
	assert len(cur_path) > 0, "在路径: {0} 中找不到 '{1}' 的表".format(path, name)

	return cur_path[0]

def check_score_int(si_list, path, cur_path, name, _min, _max):
	"""
	:检查整数分数的含范围的检查
	params：
		- si_list 要检查的数字的list， [a,b,...]
		- path 当前所在的上级路径
		- cur_path 正在处理的表名
		- name 正在处理的分数的来源
		- _min 分数的最小值（含本值）
		- _max 分数的最大值（含本值）
	"""
	for si in si_list:
		try:
			int(si)
		except:	
			raise Exception("在路径： {0} 中文件 '{1}' 的成员： '{2}' 的评分不正确，请检查该文档".format(path, cur_path, name))
		assert int(si) >= _min and int(si) <= _max, "在路径： {0} 中文件 '{1}' 的成员： '{2}' 的评分不正确，请检查该文档".format(path, cur_path, name)


def check_score_float(si_list, path, cur_path, name, _min, _max):
	"""
	:检查浮点分数的含范围的检查
	params：
		- si_list 要检查的数字的list， [a,b,...]
		- path 当前所在的上级路径
		- cur_path 正在处理的表名
		- name 正在处理的分数的来源
		- _min 分数的最小值（含本值）
		- _max 分数的最大值（含本值）
	"""
	for si in si_list:
		try:
			float(si)
		except:	
			raise Exception("在路径： {0} 中文件 '{1}' 的成员： '{2}' 的评分不正确，请检查该文档".format(path, cur_path, name))
		assert float(si) >= _min and float(si) <= _max, "在路径： {0} 中文件 '{1}' 的成员： '{2}' 的评分不正确，请检查该文档".format(path, cur_path, name)

def shuffle(foo):
	"""
	:给定一个list，打乱它的顺序，用于输出评论
	"""
	random.shuffle(foo)

def find_dep_by_name(deps, name):
	"""
	:给定一个名字，查找它所在的部门
	params:
		- deps 一个dict，其键值是部门名 {a:x1, b:x2, c:x3} a,b,c为部门名
		- name 给定的姓名
	return:
		- 部门名字的字符串，找不到则为空
	"""
	for i in deps.keys():
		for j in deps[i]:
			for k in j:
				if name == k:
					return i

	return ""
###################################################################
# :读有关主席的表                                                 #
#                                                                 #
# 变量                                                            #
#    - table1 对部长评价表                                        #
#    - table2 对其他部门评价表                                    #
#　　　　- table3 对部门评价表                                    #
# 函数                                                            #
#    - read_chair 读正主席特殊的表                                #
#    - read_chairs 读所有主席的表                                 #
# 函数依赖                                                        #
#  - read_chair   -> read_table2 -> _read_table2_helper           #
#  - read_chairs  -> read_table1                                  #
#                 -> read_table3                                  #
###################################################################
def _read_table2_helper(ws, row, col, dep, path, cur_path):
	"""
	"""
	name = ws.cell(row = row, column = col).value.strip()
	assert name in dep, "在路径 {0} 中文件 {1} 的成员 {2} 不存在，请检查配置文件".format(path,cur_path, name)

	s1 = ws.cell(row = row, column = col+5).value
	s2 = ws.cell(row = row+2, column = col+5).value
	s3 = ws.cell(row = row+4, column = col+5).value

	check_score_float([s1,s2,s3], path, cur_path, name, 0,10)
	
	return name, float(s1) + float(s2) + float(s3)

def read_table1(path, cur_path, leader_dict):
	"""
	"""
	# 读对部长评价表
	wb = load_workbook(path+cur_path)
	sheetnames = wb.get_sheet_names()  
	ws = wb.get_sheet_by_name(sheetnames[0])

	total = 0
	re_dict = {}
	for d in leader_dict:
		total += len(leader_dict[d])
		re_dict[d] = {}

	for shift in range(total):
		name = ws.cell(row = 7, column = 9+shift).value
		assert name in [j for i in leader_dict.values() for j in i], "在路径： {0} 中文件 '{1}' 的成员： '{2}' 有误，请检查该文件".format(path, cur_path, name)


		for d in leader_dict:
			sc = 0
			if name in leader_dict[d]:
				s1 = ws.cell(row = 9, column = 9+shift).value
				s2 = ws.cell(row = 11, column = 9+shift).value
				s3 = ws.cell(row = 13, column = 9+shift).value
				s4 = ws.cell(row = 15, column = 9+shift).value
				s5 = ws.cell(row = 17, column = 9+shift).value

				check_score_float([s1,s2,s3,s4,s5], path, cur_path, name,0, 10)
				
				sc = float(s1) + float(s2) + float(s3) + float(s4) + float(s5)
				break

		assert sc > 0, "在路径： {0} 中文件 '{1}' 的成员： '{2}' 不存在，请检查配置文件该部门的成员是否正确填写".format(path, cur_path, name)
		
		re_dict[d][name] = sc

	return re_dict
 
def read_table2(path, cur_path, dep):
	"""
	"""
	# 读对其他部门评价表
	wb = load_workbook(path+cur_path)
	sheetnames = wb.get_sheet_names()  
	ws = wb.get_sheet_by_name(sheetnames[0])
	
	re_dict = {}
	for shift in range(len(dep)//2):
		name, sc = _read_table2_helper(ws, 8+shift*11, 1, dep, path, cur_path)
		re_dict[name] = sc
		name, sc = _read_table2_helper(ws, 8+shift*11, 7, dep, path, cur_path)
		re_dict[name] = sc

	if len(dep)%2 == 1:
		name, sc = _read_table2_helper(ws, 8+11*(len(dep)//2), 1, dep, path, cur_path)
		re_dict[name] = sc

	return re_dict

def read_table3(path, cur_path, dep_list, exc_dict):
	"""
	"""
	# 读对部门评价表
	wb = load_workbook(path + cur_path)
	sheetnames = wb.get_sheet_names()  
	ws = wb.get_sheet_by_name(sheetnames[0])

	re_dict = {}

	for shift in range(len(dep_list)):
		name = ws.cell(row = 6, column = 7+shift).value
		assert name in dep_list, "在路径： {0} 中文件 '{1}' 的成员： '{2}' 不存在，请检查配置文件该部门的成员是否正确填写".format(path, cur_path, name)
		
		s1 = ws.cell(row = 7, column = 7+shift).value
		s2 = ws.cell(row = 8, column = 7+shift).value
		s3 = ws.cell(row = 9, column = 7+shift).value
		s4 = ws.cell(row = 10, column = 7+shift).value
		s5 = ws.cell(row = 11, column = 7+shift).value

		check_score_float([s1,s2,s3,s4,s5], path, cur_path, name, 0, 10)
		re_dict[name] = float(s1) + float(s2) + float(s3) + float(s4) + float(s5)

	e1 = ws.cell(row = 15, column = 3).value.strip()
	if e1 not in exc_dict:
		exc_dict[e1] = 0
	exc_dict[e1] += 1

	e2 = ws.cell(row = 15, column = 6).value.strip()
	if e2 not in exc_dict:	
		exc_dict[e2] = 0
	exc_dict[e2] += 1

	return (re_dict, exc_dict)

def read_chair(path, chair_name, else_dep):
	"""
	"""
	table2_path = name_to_path_by_rstr(r'.*{0}.*{1}.*'.format(chair_name, '对其他部门评价表'),path,"找不到{0}的{1},请检查该文档".format(chair_name, '对其他部门评价表'))

	table2 = read_table2(path, table2_path, else_dep)

	return table2
	
def read_chairs(path, chairs_list, dep_dict, leader_dict):
	"""
	:读所有主席的两个表，对部门评价表和对部长评价表
	params:	
		- path 放部门评价表和对部门评价表的路径
		- chirs chairs_list 主席姓名的list [a,b,...]， a,b为主席姓名
		- dep_dict 主席主管部门的dict，{a:[a1,a2...], b:[b1,b2...],...}
		                                  ,a,b为主席姓名,a1,b1,为部门名
		- leader_dict 主席主管部门的部长的dict，{a:[a1,a2...], b:[b1,b2...],...}
		                                         ,a,b为主席姓名,a1,b1,为部门名
	return:
		- (table1, table2, exc) 作为一个元组cache
		    - table1 存放所有部长评价的dict， {d1:{a:sa,b:sb}, d2:{c:sc,d:sd},...}
								d1,d2为部门名，a,b,为部长名，sa，sb为该部长的分数
		    - table3 存放对部门的分数的dict， {a:sa, b:sb,...}, 
										    a,b为部门名，sa,sb为部门分
		    - exc 存放部门推优票数的dict，但不是所有部门都有票数，{a:a1, b:b1}
										    a,b为部门名，a1,b1为推优票数
	"""
	table1 = {}
	table3 = {}
	exc = {}

	for c in chairs_list:
		table1_path = name_to_path_by_rstr(r'.*{0}.*{1}.*'.format(c, '对部长评价表'),path,"找不到{0}的{1},请检查该文档".format(c, '对部长评价表'))
		table3_path = name_to_path_by_rstr(r'.*{0}.*{1}.*'.format(c, '对部门评价表'),path,"找不到{0}的{1},请检查该文档".format(c, '对部门评价表'))

		_table1 = read_table1(path, table1_path, leader_dict[c])
		_table3, exc = read_table3(path, table3_path, dep_dict[c], exc)

		table1.update(_table1)
		table3.update(_table3)


	return (table1, table3, exc)


###################################################################
#  :读有关出勤的表                                                 #
#                                                                 #
# 函数                                                            #
#    - read_else 读其他情况加减分                                 #
#    - read_attend 读出勤表                                       #
###################################################################
def read_else(path, dep):
	"""
	"""
	# 读其他情况加减分
	r_dict = {}
	
	cur_path = '部门其他情况加减分.xlsx'
	wb = load_workbook(path + cur_path)
	sheetnames = wb.get_sheet_names()  
	ws = wb.get_sheet_by_name(sheetnames[0])
	
	dep_num = len(dep)
	for shift in range(dep_num):
		name = ws.cell(row = 2+shift, column = 1).value
		assert name in dep.keys(), "在路径： {0} 中文件 '{1}' 的成员： '{2}' 不存在，请检查配置文件该部门的成员是否正确填写".format(path, cur_path, name)
		score = ws.cell(row = 2+shift, column = 2).value
		e_score = ws.cell(row = 2+shift, column = 3).value
		
		if score is None:
			score = 0
		if e_score is None:
			e_score = 0
		check_score_float([score], path, cur_path, name, -5, 5)
		check_score_float([e_score], path, cur_path, name, -200, 200)
		r_dict[name] = score + e_score
		
	return r_dict 
	
def read_attend(path, dep_name, leader_list, member_list):
	"""
	"""
	# 读出勤分
	all_list = leader_list.copy()
	all_list.extend(member_list)

	cur_path = name_to_path(dep_name, path)

	wb = load_workbook(path + cur_path)
	sheetnames = wb.get_sheet_names()  
	ws = wb.get_sheet_by_name(sheetnames[0])

	attend_dict = {}

	d_attend_score1 = 0
	d_attend_score2 = 0

	for shift in range(len(all_list)):
		name = ws.cell(row = 3+shift*2, column = 2).value
		assert name in all_list, "在路径： {0} 中文件 '{1}' 的成员： '{2}' 不存在，请检查配置文件该部门的成员是否正确填写".format(path, cur_path, name)
		
		s1 = ws.cell(row = 3+shift*2, column = 3).value
		s2 = ws.cell(row = 3+shift*2, column = 4).value
		s3 = ws.cell(row = 3+shift*2, column = 5).value
		s4 = ws.cell(row = 3+shift*2, column = 6).value
		s5 = ws.cell(row = 3+shift*2, column = 7).value
		
		if s1 is None:
			s1 = 0
		if s2 is None:
			s2 = 0
		if s3 is None:
			s3 = 0
		if s4 is None:
			s4 = 0
		if s5 is None:
			s5 = 0

		check_score_int([s1,s2,s3,s4], path, cur_path, name, 0, 10)
		check_score_float([s5], path, cur_path, name, -5, 5)
		
		sc = 1-min(0.2*s1 + 0.4*s2 + 0.6*s3 ,1) + 0.2*s4 + s5
		attend_dict[name] = sc
		
		d_attend_score1 = d_attend_score1 - 0.2*s1 - 0.4*s2 - 0.6*s3
		d_attend_score2 = d_attend_score2 + 0.2*s4 + s5
		
	d_attend_score = 2 + max(d_attend_score1, -2)# + d_attend_score2

	return (attend_dict, d_attend_score)
	
	
###################################################################
# :读有关部长的表                                                 #
#                                                                 #
# 函数                                                            #
#    - read_leader0 读部长对干事的评价表                          #
#    - read_leader1 读部长自评表                                  #
###################################################################				
def read_leader0(path, l_list, member_list):
	"""
	"""
	# 读部长对干事的评价表
	ltom_comment_dict = {}
	ltom_score_dict = {}

	for l in l_list:
		cur_path = name_to_path(l, path)

		wb = load_workbook(path + cur_path)
		sheetnames = wb.get_sheet_names()  
		ws = wb.get_sheet_by_name(sheetnames[0])

		for shift in range(len(member_list)):
			# 部长对干事评价
			name = ws.cell(row = 22+shift, column = 1).value
			assert name in member_list, "在路径： {0} 中文件 '{1}' 的成员： '{2}' 不存在，请检查配置文件该部门的成员是否正确填写".format(path, cur_path, name)
			
			com = ws.cell(row = 22+shift, column = 2).value
			if name not in ltom_comment_dict:
				ltom_comment_dict[name] = []
			if com is None:
				com = "None"
			ltom_comment_dict[name].append(com)
			
			# 部长对干事打分
			name = ws.cell(row = 8, column = 4+shift).value.strip()
			s1 = ws.cell(row = 10, column = 4+shift).value 
			s2 = ws.cell(row = 13, column = 4+shift).value
			s3 = ws.cell(row = 16, column = 4+shift).value
						
			check_score_float([s1,s2,s3], path, cur_path, name,0,10)
			
			s = float(s1) + float(s2) + float(s3)

			if name not in ltom_score_dict:
				ltom_score_dict[name] = []
			ltom_score_dict[name].append(s)


	return (ltom_comment_dict, ltom_score_dict)

def read_leader1(path, l_list):
	"""
	"""
	ltol_comment_dict = {}
	ltol_score_dict = {}
	l_score_dict = {}

	for l in l_list:
		cur_path = name_to_path(l, path)

		wb = load_workbook(path + cur_path)
		sheetnames = wb.get_sheet_names()  
		ws = wb.get_sheet_by_name(sheetnames[0])

		# 读部长自评分
		s1 = ws.cell(row = 9, column = 4).value
		s2 = ws.cell(row = 12, column = 4).value
		s3 = ws.cell(row = 15, column = 4).value
		s4 = ws.cell(row = 18, column = 4).value
		s5 = ws.cell(row = 21, column = 4).value
		s6 = ws.cell(row = 24, column = 4).value
		s7 = ws.cell(row = 27, column = 4).value
		s8 = ws.cell(row = 30, column = 4).value

		s9 = ws.cell(row = 33, column = 4).value
		s10 = ws.cell(row = 36, column = 4).value
		s11 = ws.cell(row = 39, column = 4).value
		s12 = ws.cell(row = 42, column = 4).value

		check_score_float([s1,s2,s3,s4,s5,s6,s7,s8], path, cur_path, l,0,10)
		check_score_float([s9,s10,s11,s12], path, cur_path, l,0,5)
		
		s = float(s1)+float(s2)+float(s3)+float(s4)+float(s5)+float(s6)+float(s7)+float(s8)+float(s9)+float(s10)+float(s11)+float(s12)

		l_score_dict[l] = s

		for shift in range(len(l_list)-1):
			# 读对其他部长评价
			name = ws.cell(row = 48+shift, column = 1).value.strip()
			com = ws.cell(row = 48+shift, column = 2).value
			assert name in l_list, "在路径： {0} 中文件 '{1}' 的成员： '{2}' 不存在，请检查配置文件该部门的成员是否正确填写".format(path, cur_path, name)
			
			if name not in ltol_comment_dict:
				ltol_comment_dict[name] = []
			if com is None:
				com = "None"
			ltol_comment_dict[name].append(com)
			
			# 读对其他部长评分
			sc = ws.cell(row = 48+shift, column = 5).value

			check_score_float([sc], path, cur_path, name, 0, 10)

			if name not in ltol_score_dict:
				ltol_score_dict[name] = []
			ltol_score_dict[name].append(float(sc))

	return (ltol_comment_dict, ltol_score_dict, l_score_dict)

	
###################################################################
# :读有关干事的表                                                 #
#                                                                 #
# 函数                                                            #
#    - read_member 读部长对干事的评价表                           #
###################################################################	
def read_member(path, m_list, leader_list):
	"""
	"""
	mtol_comment_dict = {}
	mtol_score_dict = {}
	m_score_dict = {}
	exc_dict = {}

	for m in m_list:
		cur_path = name_to_path(m, path)

		wb = load_workbook(path + cur_path)
		sheetnames = wb.get_sheet_names()  
		ws = wb.get_sheet_by_name(sheetnames[0])
		
		for shift in range(len(leader_list)):
			# 读对部长的评价
			name = ws.cell(row = 42+shift, column = 1).value.strip()
			com = ws.cell(row = 42+shift, column = 2).value
			assert name in leader_list, "在路径： {0} 中文件 '{1}' member： '{2}'不存在，请检查配置文件该部门的成员是否正确填写".format(path, cur_path, name)

			if name not in mtol_comment_dict:
				mtol_comment_dict[name] = []
			if com is None:
				com = "None"
			mtol_comment_dict[name].append(com)
			
			# 读对部长的评分
			sc = ws.cell(row = 42+shift, column = 5).value

			check_score_float([sc], path, cur_path, name, 0, 10)

			if name not in mtol_score_dict:
				mtol_score_dict[name] = []
			mtol_score_dict[name].append(float(sc))

		# 读自评分
		s1 = ws.cell(row = 7, column = 4).value
		s2 = ws.cell(row = 10, column = 4).value
		s3 = ws.cell(row = 13, column = 4).value
		s4 = ws.cell(row = 16, column = 4).value
		s5 = ws.cell(row = 19, column = 4).value
		s6 = ws.cell(row = 22, column = 4).value
		s7 = ws.cell(row = 25, column = 4).value
		s8 = ws.cell(row = 28, column = 4).value
		s9 = ws.cell(row = 31, column = 4).value
		s10 = ws.cell(row = 34, column = 4).value

		check_score_float([s1,s2,s3,s4,s5,s6,s7,s8,s9,s10], path, cur_path, m, 0, 10)
		s = float(s1)+float(s2)+float(s3)+float(s4)+float(s5)+float(s6)+float(s7)+float(s8)+float(s9)+float(s10)
		m_score_dict[m] = s

		# 读干事推优
		exc = ws.cell(row = 40, column = 2).value
		if exc is None:
			continue
		exc = exc.strip()
		assert exc in m_list, "在路径： {0} 中文件 '{1}'　推优 '{2}' 不在干事名单中，请检查该文件".format(path, cur_path, exc)
		if exc not in exc_dict:
			exc_dict[exc] = 0
		exc_dict[exc] += 1


	return (mtol_comment_dict, mtol_score_dict, m_score_dict, exc_dict)

	
###################################################################
# :计算部门的总分和排序                                           #
#                                                                 #
# 函数                                                            #
#    - cal_member 计算干事的总分和排名                            #
#    - cal_leader 计算部长的总分和排名                            #
###################################################################
def cal_member(l_cache, m_cache, a_cache):
	"""
	"""
	(ltom_comment_dict, ltom_score_dict) = l_cache
	(mtol_comment_dict, _, m_score_dict, exc_dict) = m_cache
	(attend_score,_) = a_cache

	l_nums = len(mtol_comment_dict.keys())
	# 两个dict公用keys
	total_score_dict = {}
	for m in m_score_dict:
		si = 0
		for s in ltom_score_dict[m]:
			si += s
		si = si / l_nums / 6 # 所有部长平均分
		si += attend_score[m] #　出勤分
		si = si + m_score_dict[m] * 0.02 #　自评分
		if m in exc_dict:
			si = si + exc_dict[m] * 0.1 #　推优
		total_score_dict[m] = si

	sorted_total_score_dict= sorted(total_score_dict.items(), key=lambda d:d[1], reverse = True)

	return (ltom_comment_dict, total_score_dict, sorted_total_score_dict)

def cal_leader(c_cache, l_cache, m_cache, a_cache):
	"""
	"""
	(mtol_comment_dict, mtol_score_dict, m_score_dict, _) = m_cache
	(ltol_comment_dict, ltol_score_dict, l_score_dict) = l_cache
	(attend_score,_) = a_cache
	l_nums = len(c_cache)
	m_nums = len(m_score_dict)

	# 两个dict公用keys
	total_score_dict = {}
	for l in l_score_dict:
		si = 0
		for s in ltol_score_dict[l]:
			si += s

		si = 0.2 * si / (l_nums-1) #　其他部长的平均分

		s0 = 0
		for s in mtol_score_dict[l]:
			s0 += s

		si = si + 0.2 * s0 / m_nums # 干事对部长的平均分

		si += 0.1 * c_cache[l] # 主管主席对部长的评分

		si += attend_score[l] # 出勤分

		si += 0.02 * l_score_dict[l] # 自评分

		total_score_dict[l] = si

	return (mtol_comment_dict, ltol_comment_dict, total_score_dict)

	
###################################################################
# :对所有部门和部长总分进行排序，评出优秀名单                     #
#                                                                 #
# 函数                                                            #
#    - eva_dep 对所有部门总分排序，评出优秀部门                   #
#    - eva_leader 对所有部长总分排序，评出优秀部长                #
###################################################################
def eva_dep(chair, deps, s2, s3, exc, a_cache, e_cache):
	"""
	"""
	d_total_score = {}
	# 处理正主席主管部门没有其他主席评分的问题
	for d in deps[chair]:
		s2[d] = 0.6 * s3[d]
	
	for c in deps:
		for d in deps[c]:
			(_, d_attend_score) = a_cache[d]


			sc = s2[d] / 6 + 3 * s3[d] / 50 + d_attend_score
	
			if d in exc:
				sc = sc + 0.3 * exc[d]

			sc += e_cache[d]

			d_total_score[d] = sc

	sorted_d_total_score = sorted(d_total_score.items(), key=lambda d:d[1], reverse = True)

	t1_name = sorted_d_total_score[0][0]
	t1_score = sorted_d_total_score[0][1]
	t1 = (t1_name, t1_score)

	t2_name = sorted_d_total_score[1][0]
	t2_score = sorted_d_total_score[1][1]
	t2 = (t2_name, t2_score)

	return (d_total_score, sorted_d_total_score, t1, t2)

		
def eva_leader(deps, l_cache):
	"""
	"""
	all_dict = {}
	for d in deps:
		(_, _, total_score_dict) = l_cache[d]
		all_dict.update(total_score_dict)

	sorted_all_dict = sorted(all_dict.items(), key=lambda d:d[1], reverse = True)
	
	t1_name = sorted_all_dict[0][0]
	t1_score = sorted_all_dict[0][1]
	t1_dep = find_dep_by_name(deps,t1_name)
	assert t1_dep in deps.keys(), "{0} 找不到对应的部门，请检查配置文件".format(t1_name)
	t1 = (t1_name, t1_dep, t1_score)

	t2_name = sorted_all_dict[1][0]
	t2_score = sorted_all_dict[1][1]
	t2_dep = find_dep_by_name(deps,t2_name)
	assert t2_dep in deps.keys(), "{0} 找不到对应的部门，请检查配置文件".format(t2_name)
	t2 = (t2_name, t2_dep, t2_score) 

	t3_name = sorted_all_dict[2][0]
	t3_score = sorted_all_dict[2][1]
	t3_dep = find_dep_by_name(deps,t3_name)
	assert t3_dep in deps.keys(), "{0} 找不到对应的部门，请检查配置文件".format(t3_name)
	t3 = (t3_name, t3_dep, t3_score)  

	return (t1, t2, t3)

	
###################################################################
# :写反馈表                                                       #
#                                                                 #
# 函数                                                            #
#    - write_leader 写部长反馈表                                  #
#    - write_member 写干事反馈表                                  #
#    - write_all 写总的月度反馈表                                 #
###################################################################
def write_leader(d_cache, l_top3, path, dep_name, cal_m_cache, cal_l_cache):
	"""
	"""
	(mtol_comment_dict, ltol_comment_dict, total_score_dict) = cal_l_cache
	(_, _, sorted_total_score_dict) = cal_m_cache
	(d_total_score, sorted_d_total_score, d_t1, d_t2) = d_cache
	(l_t1, l_t2, l_t3) = l_top3

	wb = load_workbook('model/部长级反馈表.xlsx')
	sheetnames = wb.get_sheet_names()  
	ws = wb.get_sheet_by_name(sheetnames[0])
	
	init_borden2(ws) # 这个地方可以优化
	
	for l in mtol_comment_dict:
		shuffle(ltol_comment_dict[l])
		shuffle(mtol_comment_dict[l])

		# 部长资料
		ws.cell(row=3, column=2, value=str(dep_name))
		ws.cell(row=3, column=4, value=str(l))
		ws.cell(row=3, column=7, value="{0:.5}".format(total_score_dict[l]))

		# 优秀部门
		(t1_name, _) = d_t1
		(t2_name, _) = d_t2
		ws.cell(row=8, column=3, value=str(t1_name))
		ws.cell(row=8, column=6, value=str(t2_name))

		# 优秀部长
		(t1_name, t1_dep, _) = l_t1
		(t2_name, t2_dep, _) = l_t2
		(t3_name, t3_dep, _) = l_t3
		
		ws.cell(row=5, column=3, value=str(t1_dep))
		ws.cell(row=5, column=6, value=str(t1_name))
		ws.cell(row=6, column=3, value=str(t2_dep))
		ws.cell(row=6, column=6, value=str(t2_name))
		ws.cell(row=7, column=3, value=str(t3_dep))
		ws.cell(row=7, column=6, value=str(t3_name))

		# 本部门干事排名
		for shift in range(len(sorted_total_score_dict)):
			ws.cell(row=10+shift, column=10, value=str(sorted_total_score_dict[shift][0]))
			ws.cell(row=10+shift, column=12, value="{0:.5}".format(sorted_total_score_dict[shift][1]))

		# 部门总排名
		for shift in range(len(sorted_d_total_score)):
			ws.cell(row=10+shift, column=15, value=str(sorted_d_total_score[shift][0]))
			ws.cell(row=10+shift, column=17, value="{0:.5}".format(sorted_d_total_score[shift][1]))

		# 其他部长对部长的评价
		str0 = ''
		for c in ltol_comment_dict[l]:
			str0 = "{0}\n\n{1}".format(str0, c)
		ws.cell(row=10, column=3, value=str0)

		# 干事对部长的评价
		str1 = ''
		for c in mtol_comment_dict[l]:
			str1 = "{0}\n\n{1}".format(str1, c)
		ws.cell(row=18, column=3, value=str1)


		wb.save(path+l+'部长级反馈表.xlsx')

def write_member(d_cache, l_top3, path, dep_name, cal_cache):
	"""
	"""
	(ltom_comment_dict, total_score_dict, sorted_total_score_dict) = cal_cache
	(d_total_score, _, d_t1, d_t2) = d_cache
	(l_t1, l_t2, l_t3) = l_top3

	wb = load_workbook('model/干事反馈表.xlsx')
	sheetnames = wb.get_sheet_names()  
	ws = wb.get_sheet_by_name(sheetnames[0])
		
	init_borden1(ws) # 这个地方可以优化
		
	for m in ltom_comment_dict:
		shuffle(ltom_comment_dict[m])

		# 干事资料	
		ws.cell(row=3, column=3, value=str(dep_name))
		ws.cell(row=3, column=7, value=str(m))
		ws.cell(row=4, column=3, value="{0:.5}".format(d_total_score[dep_name]))
		ws.cell(row=4, column=7, value="{0:.5}".format(total_score_dict[m]))

		# 优秀部门
		(t1_name, _) = d_t1
		(t2_name, _) = d_t2
		ws.cell(row=6, column=3, value=str(t1_name))
		ws.cell(row=6, column=6, value=str(t2_name))

		# 优秀部长
		(t1_name, t1_dep, _) = l_t1
		(t2_name, t2_dep, _) = l_t2
		(t3_name, t3_dep, _) = l_t3

		ws.cell(row=7, column=3, value=str(t1_dep))
		ws.cell(row=7, column=6, value=str(t1_name))
		ws.cell(row=8, column=3, value=str(t2_dep))
		ws.cell(row=8, column=6, value=str(t2_name))
		ws.cell(row=9, column=3, value=str(t3_dep))
		ws.cell(row=9, column=6, value=str(t3_name))

		# 写部门前三	
		ws.cell(row=10, column=3, value=str(sorted_total_score_dict[0][0]))
		ws.cell(row=10, column=6, value="{0:.5}".format(sorted_total_score_dict[0][1]))
		ws.cell(row=11, column=3, value=str(sorted_total_score_dict[1][0]))
		ws.cell(row=11, column=6, value="{0:.5}".format(sorted_total_score_dict[1][1]))
		ws.cell(row=12, column=3, value=str(sorted_total_score_dict[2][0]))
		ws.cell(row=12, column=6, value="{0:.5}".format(sorted_total_score_dict[2][1]))

		# 写部长对干事的评论
		str0 = ''
		for c in ltom_comment_dict[m]:
			str0 = "{0}\n\n{1}".format(str0, c)
		ws.cell(row=14, column=3, value=str0)

		wb.save(path+m+'干事反馈表.xlsx')

def write_all(deps, d_cache, l_top3, path, cal_cache):
	"""
	"""
	(_, _, d_t1, d_t2) = d_cache
	(l_t1, l_t2, l_t3) = l_top3

	wb = load_workbook('model/绩效考核反馈表.xlsx')
	sheetnames = wb.get_sheet_names()  
	ws = wb.get_sheet_by_name(sheetnames[0])

	init_borden3(ws, len(deps)) # 这个地方可以优化

	# 优秀部门
	(t1_name, t1_score) = d_t1
	(t2_name, t2_score) = d_t2
	ws.cell(row=6, column=1, value=str(t1_name))
	ws.cell(row=6, column=4, value="{0:.5}".format(t1_score))
	ws.cell(row=9, column=1, value=str(t2_name))
	ws.cell(row=9, column=4, value="{0:.5}".format(t2_score))

	# 优秀部长
	(t1_name, t1_dep, t1_score) = l_t1
	(t2_name, t2_dep, t2_score) = l_t2
	(t3_name, t3_dep, t3_score) = l_t3

	ws.cell(row=6, column=9, value=str(t1_dep))
	ws.cell(row=6, column=11, value=str(t1_name))
	ws.cell(row=6, column=13, value="{0:.5}".format(t1_score))
	ws.cell(row=8, column=9, value=str(t2_dep))
	ws.cell(row=8, column=11, value=str(t2_name))
	ws.cell(row=8, column=13, value="{0:.5}".format(t2_score))
	ws.cell(row=10, column=9, value=str(t3_dep))
	ws.cell(row=10, column=11, value=str(t3_name))
	ws.cell(row=10, column=13, value="{0:.5}".format(t3_score))

	# 各部门前三
	for shift in range(len(deps)//5):
		for shift2 in range(5):
			name = ws.cell(row = 13+shift*9, column=1+shift2*3).value.strip()
			assert name in deps.keys(), "model文件 {0} 的成员： {1} 不存在，请检查该文件".format('绩效考核反馈表.xlsx',name)
			(_, _, sorted_total_score_dict) = cal_cache[name]
			ws.cell(row=15+shift*9, column=1+shift2*3, value=str(sorted_total_score_dict[0][0]))
			ws.cell(row=15+shift*9, column=2+shift2*3, value="{0:.5}".format(sorted_total_score_dict[0][1]))
			ws.cell(row=17+shift*9, column=1+shift2*3, value=str(sorted_total_score_dict[1][0]))
			ws.cell(row=17+shift*9, column=2+shift2*3, value="{0:.5}".format(sorted_total_score_dict[1][1]))
			ws.cell(row=19+shift*9, column=1+shift2*3, value=str(sorted_total_score_dict[2][0]))
			ws.cell(row=19+shift*9, column=2+shift2*3, value="{0:.5}".format(sorted_total_score_dict[2][1]))

	for shift in range(len(deps)%5):
		name = ws.cell(row = 13+(len(deps)//5)*9, column=1+shift*3).value.strip()
		assert name in deps.keys(), "model文件 {0} 的成员： {1} 不存在，请检查该文件".format('绩效考核反馈表.xlsx',name)
		(_, _, sorted_total_score_dict) = cal_cache[name]
		ws.cell(row=15+(len(deps)//5)*9, column=1+shift*3, value=str(sorted_total_score_dict[0][0]))
		ws.cell(row=15+(len(deps)//5)*9, column=2+shift*3, value="{0:.5}".format(sorted_total_score_dict[0][1]))
		ws.cell(row=17+(len(deps)//5)*9, column=1+shift*3, value=str(sorted_total_score_dict[1][0]))
		ws.cell(row=17+(len(deps)//5)*9, column=2+shift*3, value="{0:.5}".format(sorted_total_score_dict[1][1]))
		ws.cell(row=19+(len(deps)//5)*9, column=1+shift*3, value=str(sorted_total_score_dict[2][0]))
		ws.cell(row=19+(len(deps)//5)*9, column=2+shift*3, value="{0:.5}".format(sorted_total_score_dict[2][1]))	

	wb.save(path+'绩效考核反馈表.xlsx')
