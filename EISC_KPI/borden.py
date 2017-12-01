from openpyxl.styles import Border,Side

def init_borden1(ws):
	"""
	"""
	left, right, top, bottom = [Side(style='medium', color='000000')]*4
	border = Border(left=left, right=right, top=top, bottom=bottom)
	
	ws.cell(row=1, column=8, value="").border=border
	ws.cell(row=2, column=8, value="").border=border
	
	str0 = ws.cell(row = 3, column = 1).value.strip()
	str1 = ws.cell(row = 3, column = 5).value.strip()
	str2 = ws.cell(row = 4, column = 1).value.strip()
	str3 = ws.cell(row = 4, column = 5).value.strip()
	str4 = ws.cell(row = 6, column = 1).value.strip()
	str5 = ws.cell(row = 7, column = 1).value.strip()
	str6 = ws.cell(row = 10, column = 1).value.strip()
	str7 = ws.cell(row = 14, column = 1).value.strip()

	for i in range(2,30):
		for j in range(1, 9):
			ws.cell(row=i, column=j, value="").border=border
	
	ws.cell(row=3, column=1, value=str0)
	ws.cell(row=3, column=5, value=str1)
	ws.cell(row=4, column=1, value=str2)
	ws.cell(row=4, column=5, value=str3)
	ws.cell(row=6, column=1, value=str4)
	ws.cell(row=7, column=1, value=str5)
	ws.cell(row=10, column=1, value=str6)
	ws.cell(row=14, column=1, value=str7)

def init_borden2(ws):
	"""
	"""
	left, right, top, bottom = [Side(style='medium', color='000000')]*4
	border = Border(left=left, right=right, top=top, bottom=bottom)
	
	ws.cell(row=1, column=8, value="").border=border
	ws.cell(row=2, column=8, value="").border=border
	
	str0 = ws.cell(row = 3, column = 1).value.strip()
	str1 = ws.cell(row = 3, column = 3).value.strip()
	str2 = ws.cell(row = 3, column = 5).value.strip()
	str3 = ws.cell(row = 5, column = 1).value.strip()
	str4 = ws.cell(row = 8, column = 1).value.strip()
	str5 = ws.cell(row = 10, column = 1).value.strip()
	str6 = ws.cell(row = 18, column = 1).value.strip()

	for i in range(2,43):
			for j in range(1, 9):
				ws.cell(row=i, column=j, value="").border=border

	ws.cell(row=3, column=1, value=str0)
	ws.cell(row=3, column=3, value=str1)
	ws.cell(row=3, column=5, value=str2)
	ws.cell(row=5, column=1, value=str3)
	ws.cell(row=8, column=1, value=str4)
	ws.cell(row=10, column=1, value=str5)
	ws.cell(row=18, column=1, value=str6)

	str0 = ws.cell(row = 7, column = 10).value.strip()
	for i in range(7,25):
			for j in range(10, 14):
				ws.cell(row=i, column=j, value="").border=border
	ws.cell(row=7, column=10, value=str0)

	str0 = ws.cell(row = 7, column = 15).value.strip()
	for i in range(7,20):
			for j in range(15, 19):
				ws.cell(row=i, column=j, value="").border=border
	ws.cell(row=7, column=15, value=str0)
			
def init_borden3(ws, n):
	"""
	"""
	left, right, top, bottom = [Side(style='thin', color='000000')]*4
	border = Border(left=left, right=right, top=top, bottom=bottom)

	for i in range(1,15):
		ws.cell(row=2, column=i, value="").border=border
	
	ws.cell(row=1, column=14, value="").border=border
	ws.cell(row=2, column=14, value="").border=border

	str0 = ws.cell(row = 4, column = 1).value.strip()
	for i in range(4,12):
			for j in range(1, 7):
				ws.cell(row=i, column=j, value="").border=border
	ws.cell(row=4, column=1, value=str0)
	
	str0 = ws.cell(row = 4, column = 9).value.strip()
	for i in range(4,12):
			for j in range(9, 15):
				ws.cell(row=i, column=j, value="").border=border
	ws.cell(row=4, column=9, value=str0)
	
	for i in range(n//5):
		for j in range(5):
			
			str0 = ws.cell(row = 13+9*i, column = 1+3*j).value.strip()
			for k in range(8):
				for l in range(2):
					ws.cell(row=k+13+9*i, column=1+l+3*j, value="").border=border
			ws.cell(row=13+9*i, column=1+3*j, value=str0)
			
	for i in range(n%5):
		str0 = ws.cell(row = 13+9*(n//5), column = 1+3*i).value.strip()
		for k in range(8):
				for l in range(2):
					ws.cell(row=k+13+9*(n//5), column=1+l+3*i, value="").border=border
		ws.cell(row=13+9*(n//5), column=1+3*i, value=str0)
			

		ws.cell(row=13+9*(n//5), column=1+3*j, value=str0)
			

